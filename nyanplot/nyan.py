import numpy as np
import matplotlib.pyplot as plt
from decimal import*
from mpmath import *
from matplotlib.ticker import (MultipleLocator, AutoMinorLocator)
#from decimal import *
import math
import openpyxl
mp.dps = 10#10桁までの精度
sheetn = []
std_y = np.zeros([30])
for i in range(30):
    sheetn +=["label{}".format(i)]
al = ["A","B","C","D","E","F","G","H","I"]

def read_sheet(sheetname,filename):
    file = openpyxl.load_workbook(filename)
    sheet = file.get_sheet_by_name(sheetname)
    nrow=sheet.max_row-1
    ncol=sheet.max_column
    #print(nrow,ncol)
    data = np.zeros([ncol,nrow])
    tit = {}#タイトル
    titlist=[]
    for i in range(ncol): 
        tit.setdefault(sheet["{0}1".format(al[i])].value,i)
        titlist+=[str(sheet["{0}1".format(al[i])].value)]
        #print(type(sheet["{0}1".format(al[i])].value))
    for i in range(nrow):
        for l in range(ncol):
            data[l][i]=sheet["{0}{1}".format(al[l],i+2)].value
    return(nrow,ncol,data)

def read_file(filename="filename"):
    file = openpyxl.load_workbook(filename)
    sheetn = file.get_sheet_names()
    print("シート名:",sheetn)
    N = len(sheetn)
    data = np.zeros(0)
    for i in range(N):
        num_row,num_colum,inputdata = read_sheet(sheetn[i],filename)
        data = np.append(data,inputdata)
    data = data.reshape([N,num_colum,num_row])
    return (data,sheetn)

def max_sig(x):#小数何桁目までか 小数を含まないとおかしくなる4桁までしか正しくできない
    if x == 0:
        return 0

    # 指数表記に変換して有効数字を考慮
    exponent = 1 + int(math.floor(math.log10(abs(x))))
    return exponent

def min_sig(x):#小数何桁目までか 小数を含まないとおかしくなる4桁までしか正しくできない
    if x == 0:
        return 0

    # 指数表記に変換して有効数字を考慮
    length = len(str(abs(x)).replace(".",""))
    exponent = -1*length +1+ int(math.floor(math.log10(abs(x))))
    return exponent

def round_to_sigfigs(number, sigfigs):
    if number == 0:
        return 0

    # 指数表記に変換して有効数字を考慮
    exponent = sigfigs - 1 - int(math.floor(math.log10(abs(number))))
    rounded_number = round(number, exponent)

    # 科学的表記になっている場合、有効数字の桁数を考慮して小数点以下の桁数を修正
    if 'e' in f'{rounded_number}':
        decimal_places = sigfigs - 1
        return f'{rounded_number:.{decimal_places}e}'
    else:
        return rounded_number
    
def std(ndata):#standard devision
    return np.std(ndata)/len(ndata)

def aggregate(data,n):#データ,重複度
    xydata=np.zeros(0)
    std_ydata=np.zeros(0)
    for j in range(len(data)):
        xdata = np.zeros(0)
        ydata = np.zeros(0)
        for i in range(len(data[j][0])//n):
            std_y = round_to_sigfigs(std(data[j][1][n*i:n*i+n]),2)
            ydata=np.append(ydata,np.average(data[j][1][n*i:n*i+n]))
            std_ydata=np.append(std_ydata,std_y)
            #ydata=np.append(ydata,round_to_sigfigs(np.average(data[j][1][n*i:n*i+n]),max_sig(std_y)-3))
            xdata=np.append(xdata,data[j][0][n*i])
            #print("i:",i,"std_y;",std_y)
        xydata = np.append(xydata,xdata)
        xydata = np.append(xydata,ydata)
    #print("xdata; ",xdata)
    #print("ydata; ",ydata)
    return xydata.reshape(len(data),2,len(data[j][0])//n),std_ydata.reshape(len(data),len(data[j][0])//n)

def plot(data,graph_title="title",x_label="x_label",y_label="y_label",xlog = False,ylog = False,label=sheetn, errorbar = False,std_y = std_y):
    plt.show()
    plt.rcParams["font.family"] = "meiryo"
    fig,ax= plt.subplots()
    point_num = 1000
    x = np.linspace(0,1000,10000)
    ndata = len(data[0,1])
    demention = len(np.shape(data))
    #ticksの設定
    yrange = np.max(data[:,1,:])-np.min(data[:,1,:])
    xrange = np.max(data[:,0,:])-np.min(data[:,0,:])
    xdigit =  int(math.floor(math.log10(abs(xrange/ndata))))
    ydigit =  int(math.floor(math.log10(abs(yrange/ndata))))

    if demention ==2:
        data = data.reshape(1,np.shape(data)[0],np.shape(data)[1])
    #表示範囲の設定
    plt.ylim(np.min(data[:,1,:])-yrange/ndata,np.max(data[:,1,:])+yrange/ndata)
    plt.xlim(np.min(data[:,0,:])-xrange/ndata,np.max(data[:,0,:])+xrange/ndata)

    
    x = np.linspace(np.min(data[:,0,:])-xrange/ndata,np.max(data[:,0,:])+xrange/ndata,point_num)
    A = np.zeros(0)
    B = np.zeros(0)
    for i in range(len(data)):
        plt.scatter(data[i][0],data[i][1],label=label[i])
        a,b = np.polyfit(data[i][0],data[i][1],1)
        A = np.append(A,a)
        B = np.append(B,b)
        print("sheet = {0} \n a={1} \n b = {2} \n".format(sheetn[i],a,b))
        print("-"*10)
        if errorbar == True:
            plt.errorbar(data[i][0], data[i][1], yerr = std_y[i],capsize=5, fmt="o",markersize=5,ecolor="black")#,color='r',markeredgecolor = "r")
        plt.plot(x,a*x+b,label = sheetn[i])
    print()
    #plt.plot()
    if xlog == True:
        plt.yscale("log")#対数表示
    if ylog == True:
        plt.xscale("log")#対数表示
    plt.legend()
    plt.ylabel(y_label,size=16,fontname="MS Gothic")
    plt.xlabel(x_label,size=16,fontname="MS Gothic")
    ax.xaxis.set_minor_locator(MultipleLocator(10**xdigit))
    ax.yaxis.set_minor_locator(MultipleLocator(10**ydigit))
    #ax.set_yticks(np.arange(np.min(data[:,1,:])-yrange/ndata, np.max(data[:,1,:])+yrange/ndata,(np.max(data[:,1,:])-np.min(data[:,1,:]))/30), minor=True)
    plt.title(graph_title,fontname="MS Gothic",size=20,loc="center",y = -0.25)
    plt.rcParams["xtick.direction"] = "in"      # 目盛り線の向き、内側"in"か外側"out"かその両方"inout"か
    plt.rcParams["ytick.direction"] = "in"      # 目盛り線の向き、内側"in"か外側"out"かその両方"inout"か
    plt.rcParams["xtick.top"] = True            # 上部に目盛り線を描くかどうか
    plt.rcParams["xtick.bottom"] = True         # 下部に目盛り線を描くかどうか
    plt.rcParams["ytick.left"] = True           # 左部に目盛り線を描くかどうか
    plt.rcParams["ytick.right"] = True
    plt.show()
